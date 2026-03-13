"""AI Business Assistant — Document processing, chat, preview & post engine."""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from pydantic import BaseModel
from typing import Optional, List, Any
from datetime import datetime, timezone
from database import db
from deps import get_current_user, resolve_company_id
from models import UserRole
from odoo_accounting.engine import create_invoice_move, register_payment
import uuid
import os
import json
import re
import logging
import base64
import io

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/ai-assistant")

EMERGENT_KEY_ENV = "EMERGENT_LLM_KEY"

# ============ HELPERS ============

async def _cid(user, company_id=None):
    cid = await resolve_company_id(user['user_id'], user['role'], company_id)
    if not cid:
        c = await db.companies.find_one({"status": "active"}, {"_id": 0, "id": 1})
        cid = c["id"] if c else None
    return cid


async def _context(cid):
    """Build rich business context for AI prompts."""
    accounts = await db.odoo_accounts.find({"company_id": cid}, {"_id": 0, "id": 1, "code": 1, "name": 1, "account_type": 1}).sort("code", 1).to_list(500)
    partners = await db.odoo_partners.find({"company_id": cid}, {"_id": 0, "id": 1, "name": 1, "partner_type": 1, "gstin": 1}).to_list(500)
    journals = await db.odoo_journals.find({"company_id": cid}, {"_id": 0, "id": 1, "name": 1, "code": 1, "journal_type": 1}).to_list(50)
    products = await db.products.find({"company_id": cid}, {"_id": 0, "id": 1, "name": 1, "sku": 1, "sale_price": 1, "cost_price": 1, "quantity_on_hand": 1, "uom": 1}).to_list(500)
    warehouses = await db.stock_warehouses.find({"company_id": cid}, {"_id": 0, "id": 1, "name": 1}).to_list(50)
    locations = await db.stock_locations.find({"company_id": cid}, {"_id": 0, "id": 1, "name": 1, "location_type": 1}).to_list(100)
    # Correction mappings for smart learning
    mappings = await db.ai_correction_mappings.find({"company_id": cid}, {"_id": 0}).to_list(200)
    return {
        "accounts": accounts, "partners": partners, "journals": journals,
        "products": products, "warehouses": warehouses, "locations": locations,
        "mappings": mappings,
    }


def _fmt_accounts(accts):
    return "\n".join([f"- {a['code']} {a['name']} (id:{a['id']}, type:{a['account_type']})" for a in accts[:80]])


def _fmt_partners(ps):
    return "\n".join([f"- {p['name']} (id:{p['id']}, type:{p.get('partner_type','')}, gstin:{p.get('gstin','')})" for p in ps[:60]])


def _fmt_products(ps):
    return "\n".join([f"- {p['name']} (id:{p['id']}, sku:{p.get('sku','')}, qty:{p.get('quantity_on_hand',0)}, cost:{p.get('cost_price',0)})" for p in ps[:80]])


def _fmt_journals(js):
    return "\n".join([f"- {j['name']} ({j['code']}, type:{j['journal_type']}, id:{j['id']})" for j in js])


def _fmt_warehouses(ws):
    return "\n".join([f"- {w['name']} (id:{w['id']})" for w in ws]) or "No warehouses"


def _fmt_mappings(ms):
    if not ms:
        return "No correction mappings yet."
    return "\n".join([f"- '{m['original']}' → '{m['corrected']}' (field:{m.get('field','')})" for m in ms[:30]])


async def _call_llm(system_prompt, user_prompt, model="gpt-4o-mini"):
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    key = os.environ.get(EMERGENT_KEY_ENV)
    if not key:
        raise HTTPException(status_code=500, detail="AI service not configured")
    chat = LlmChat(api_key=key, session_id=f"ai-asst-{uuid.uuid4().hex[:8]}", system_message=system_prompt)
    chat = chat.with_model("openai", model)
    return await chat.send_message(UserMessage(text=user_prompt))


async def _call_vision(system_prompt, user_prompt, image_b64):
    from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
    key = os.environ.get(EMERGENT_KEY_ENV)
    if not key:
        raise HTTPException(status_code=500, detail="AI service not configured")
    chat = LlmChat(api_key=key, session_id=f"ai-vision-{uuid.uuid4().hex[:8]}", system_message=system_prompt)
    chat = chat.with_model("openai", "gpt-4o")
    msg = UserMessage(text=user_prompt, file_contents=[ImageContent(image_base64=image_b64)])
    return await chat.send_message(msg)


def _parse_json(text):
    """Parse JSON from AI response — handles single object or array."""
    m = re.search(r'```(?:json)?\s*([\s\S]*?)```', text)
    s = m.group(1).strip() if m else text.strip()
    return json.loads(s)


def _ensure_batch(data):
    """Normalize AI response to always be a list of entries."""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        # Check if it has batch_entries
        if "batch_entries" in data and isinstance(data["batch_entries"], list):
            return data["batch_entries"]
        # Single entry
        if data.get("action_type"):
            return [data]
    return [data]


def _extract_file_text(content: bytes, filename: str) -> str:
    """Extract text from PDF, Excel, CSV files with structure preservation."""
    ext = os.path.splitext(filename)[1].lower()
    if ext == '.csv':
        text = content.decode('utf-8', errors='replace')
        lines = text.strip().split('\n')
        numbered = []
        for i, line in enumerate(lines):
            prefix = "HEADER:" if i == 0 else f"ROW {i}:"
            numbered.append(f"{prefix} {line}")
        return "\n".join(numbered)
    elif ext in ('.xlsx', '.xls'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"\n=== Sheet: {ws.title} (rows: {ws.max_row}) ===")
            headers = []
            row_count = 0
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = [str(c) if c is not None else "" for c in row]
                if not any(c.strip() for c in cells):
                    continue
                if row_idx == 1:
                    headers = cells
                    lines.append(f"COLUMNS: {' | '.join(cells)}")
                else:
                    row_count += 1
                    if headers and len(headers) == len(cells):
                        pairs = [f"{h}={v}" for h, v in zip(headers, cells) if v.strip()]
                        lines.append(f"ROW {row_idx}: {', '.join(pairs)}")
                    else:
                        lines.append(f"ROW {row_idx}: {' | '.join(cells)}")
            if row_count == 0:
                lines.append("  (no data rows)")
        return "\n".join(lines)
    elif ext == '.pdf':
        import pdfplumber
        text_parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        text_parts.append("\t".join([str(c) if c else "" for c in row]))
        return "\n".join(text_parts)
    else:
        return content.decode('utf-8', errors='replace')


# ============ MODELS ============

class ChatMessage(BaseModel):
    message: str
    company_id: Optional[str] = None
    session_id: Optional[str] = None

class ApproveRequest(BaseModel):
    pending_id: str
    entries: Optional[dict] = None  # User-edited entries

class BatchApproveRequest(BaseModel):
    pending_ids: List[str]

class RejectRequest(BaseModel):
    pending_id: str

class CorrectionMapping(BaseModel):
    original: str
    corrected: str
    field: Optional[str] = "name"
    company_id: Optional[str] = None


# ============ 1. CHAT ENDPOINT ============

@router.post("/chat")
async def ai_chat(req: ChatMessage, current_user: dict = Depends(get_current_user)):
    """Main chat endpoint. Handles natural language questions, returns answers with data."""
    if current_user['role'] == UserRole.GROUND_STAFF:
        raise HTTPException(status_code=403, detail="Access denied")

    cid = await _cid(current_user, req.company_id)
    ctx = await _context(cid)

    # Get recent financial summary
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1).strftime("%Y-%m-%d")
    today = now.strftime("%Y-%m-%d")

    invoices_today = await db.odoo_moves.count_documents({"company_id": cid, "date": today, "move_type": {"$in": ["out_invoice", "in_invoice"]}})
    total_sales = 0
    async for inv in db.odoo_moves.find({"company_id": cid, "move_type": "out_invoice", "state": "posted", "date": {"$gte": month_start}}, {"amount_total": 1, "_id": 0}):
        total_sales += inv.get("amount_total", 0)
    total_purchases = 0
    async for inv in db.odoo_moves.find({"company_id": cid, "move_type": "in_invoice", "state": "posted", "date": {"$gte": month_start}}, {"amount_total": 1, "_id": 0}):
        total_purchases += inv.get("amount_total", 0)

    # Outstanding receivables
    outstanding = 0
    async for inv in db.odoo_moves.find({"company_id": cid, "move_type": "out_invoice", "state": "posted", "amount_residual": {"$gt": 0}}, {"amount_residual": 1, "_id": 0}):
        outstanding += inv.get("amount_residual", 0)

    system_prompt = f"""You are an AI Business Assistant for an Indian multi-business ERP system.
You help Directors and Managers with accounting, inventory, and financial queries.

CURRENT BUSINESS CONTEXT:
- Company ID: {cid}
- Date: {today}
- Month Sales: INR {total_sales:,.2f}
- Month Purchases: INR {total_purchases:,.2f}
- Invoices Today: {invoices_today}
- Outstanding Receivables: INR {outstanding:,.2f}

CHART OF ACCOUNTS:
{_fmt_accounts(ctx['accounts'])}

PARTNERS:
{_fmt_partners(ctx['partners'])}

PRODUCTS/INVENTORY:
{_fmt_products(ctx['products'])}

JOURNALS:
{_fmt_journals(ctx['journals'])}

WAREHOUSES:
{_fmt_warehouses(ctx['warehouses'])}

LEARNED CORRECTIONS:
{_fmt_mappings(ctx['mappings'])}

RULES:
- Answer business questions with data from context
- If asked to create entries, return JSON with action_type
- NEVER auto-post. Always return preview data
- For financial queries, include numbers and insights
- Format currency as INR with commas (Indian format)
- If the user asks for MULTIPLE entries (e.g. "create 5 journal entries" or gives a list), return a JSON ARRAY of entry objects
- Each entry in the array must be a complete entry object with its own action_type, accounting_entries, etc.

RESPONSE FORMAT for questions: Plain text with data
RESPONSE FORMAT for SINGLE entry request: JSON object with:
```json
{{
  "action_type": "journal_entry|invoice|payment|inventory_adjustment|stock_move",
  "description": "what this entry does",
  "accounting_entries": [
    {{"account_code": "...", "account_name": "...", "account_id": "...", "debit": 0, "credit": 0, "description": "..."}}
  ],
  "inventory_entries": [
    {{"product_name": "...", "product_id": "...", "quantity_change": 0, "warehouse": "...", "warehouse_id": "..."}}
  ],
  "invoice_data": null,
  "payment_data": null,
  "gst": {{"cgst": 0, "sgst": 0, "igst": 0}},
  "total_amount": 0,
  "partner_name": "...",
  "partner_id": "..."
}}
```

RESPONSE FORMAT for MULTIPLE entries: JSON object with batch_entries array:
```json
{{
  "batch_entries": [
    {{"action_type": "...", "description": "...", "accounting_entries": [...], "inventory_entries": [...], "gst": {{...}}, "total_amount": 0, "partner_name": "...", "partner_id": "..."}},
    ...more entries...
  ]
}}
```"""

    try:
        resp = await _call_llm(system_prompt, req.message, "gpt-4o")

        # Try to parse as JSON action(s)
        try:
            data = _parse_json(resp)
            entries_list = _ensure_batch(data)

            if entries_list and isinstance(entries_list[0], dict) and entries_list[0].get("action_type"):
                if len(entries_list) == 1:
                    # Single entry — same as before
                    entry = entries_list[0]
                    pending_id = str(uuid.uuid4())
                    await db.ai_pending_entries.insert_one({
                        "_id": pending_id,
                        "id": pending_id,
                        "company_id": cid,
                        "user_id": current_user["user_id"],
                        "user_name": (await db.users.find_one({"id": current_user["user_id"]}, {"name": 1, "_id": 0}) or {}).get("name", ""),
                        "action_type": entry["action_type"],
                        "entries": entry,
                        "original_message": req.message,
                        "status": "pending",
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "source": "chat",
                    })
                    return {
                        "type": "preview",
                        "pending_id": pending_id,
                        "message": entry.get("description", "AI has prepared the following entries for your review."),
                        "entries": entry,
                    }
                else:
                    # Multiple entries — batch mode
                    batch_ids = []
                    user_name = (await db.users.find_one({"id": current_user["user_id"]}, {"name": 1, "_id": 0}) or {}).get("name", "")
                    now_iso = datetime.now(timezone.utc).isoformat()
                    batch_ref = str(uuid.uuid4())[:8]
                    for idx, entry in enumerate(entries_list):
                        pending_id = str(uuid.uuid4())
                        await db.ai_pending_entries.insert_one({
                            "_id": pending_id,
                            "id": pending_id,
                            "company_id": cid,
                            "user_id": current_user["user_id"],
                            "user_name": user_name,
                            "action_type": entry.get("action_type", "journal_entry"),
                            "entries": entry,
                            "original_message": req.message,
                            "status": "pending",
                            "created_at": now_iso,
                            "source": "chat",
                            "batch_ref": batch_ref,
                            "batch_index": idx,
                        })
                        batch_ids.append({"pending_id": pending_id, "entries": entry})
                    return {
                        "type": "batch_preview",
                        "batch_ref": batch_ref,
                        "message": f"AI has prepared {len(batch_ids)} entries for your review.",
                        "batch": batch_ids,
                        "total_count": len(batch_ids),
                    }
        except (json.JSONDecodeError, ValueError):
            pass

        # Plain text answer
        return {"type": "answer", "message": resp}

    except Exception as e:
        logger.error(f"AI chat error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 2. FILE UPLOAD & ANALYZE ============

@router.post("/upload")
async def ai_upload_file(
    file: UploadFile = File(...),
    message: str = Form(""),
    company_id: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user),
):
    """Upload a file (PDF, Excel, CSV, Image) for AI analysis and entry creation."""
    if current_user['role'] == UserRole.GROUND_STAFF:
        raise HTTPException(status_code=403, detail="Access denied")

    content = await file.read()
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 15MB)")

    cid = await _cid(current_user, company_id)
    ctx = await _context(cid)
    fname = file.filename or "upload"
    ext = os.path.splitext(fname)[1].lower()
    mime = file.content_type or ""

    # Save file
    saved_name = f"ai_upload_{uuid.uuid4().hex[:12]}{ext}"
    saved_path = os.path.join("/app/uploads", saved_name)
    with open(saved_path, "wb") as f:
        f.write(content)
    file_url = f"/api/files/{saved_name}"

    is_image = ext in ('.jpg', '.jpeg', '.png', '.webp') or mime.startswith('image/')

    context_block = f"""CHART OF ACCOUNTS:\n{_fmt_accounts(ctx['accounts'])}
PARTNERS:\n{_fmt_partners(ctx['partners'])}
PRODUCTS/INVENTORY:\n{_fmt_products(ctx['products'])}
JOURNALS:\n{_fmt_journals(ctx['journals'])}
WAREHOUSES:\n{_fmt_warehouses(ctx['warehouses'])}
LEARNED CORRECTIONS:\n{_fmt_mappings(ctx['mappings'])}"""

    system_prompt = f"""You are an AI document processing engine for an Indian multi-business ERP.
Analyze the uploaded document and extract ALL structured data.

{context_block}

STEP 1: Detect document type (purchase_invoice, sales_invoice, bank_statement, stock_report, expense_bill, delivery_challan, purchase_order, transport_receipt, other)
STEP 2: Extract all data fields — process EVERY row/line item in the document
STEP 3: Map to accounting entries + inventory entries + GST

IMPORTANT: If the document contains MULTIPLE transactions/invoices/entries (e.g. a spreadsheet with many rows, a bank statement with many transactions), you MUST create a SEPARATE entry for each transaction.

For a SINGLE transaction, respond with:
```json
{{
  "document_type": "...",
  "confidence": 0.95,
  "extracted_data": {{
    "vendor_name": "...",
    "customer_name": "...",
    "document_number": "...",
    "document_date": "YYYY-MM-DD",
    "line_items": [
      {{
        "description": "...",
        "hsn_code": "...",
        "quantity": 0,
        "unit": "...",
        "rate": 0,
        "amount": 0,
        "gst_rate": 0
      }}
    ],
    "subtotal": 0,
    "gst_details": {{"cgst": 0, "sgst": 0, "igst": 0, "total_tax": 0}},
    "grand_total": 0,
    "payment_terms": "...",
    "notes": "..."
  }},
  "action_type": "invoice|inventory_adjustment|stock_move|reconciliation|journal_entry",
  "description": "human-readable description of what was detected",
  "accounting_entries": [
    {{"account_code": "...", "account_name": "...", "account_id": "...", "debit": 0, "credit": 0, "description": "..."}}
  ],
  "inventory_entries": [
    {{"product_name": "...", "product_id": "...", "quantity_change": 0, "warehouse": "...", "warehouse_id": "...", "unit": "..."}}
  ],
  "gst": {{"cgst": 0, "sgst": 0, "igst": 0}},
  "total_amount": 0,
  "partner_name": "...",
  "partner_id": "..."
}}
```

For MULTIPLE transactions/entries from the document, respond with:
```json
{{
  "batch_entries": [
    {{
      "document_type": "...",
      "confidence": 0.95,
      "action_type": "...",
      "description": "Entry 1 — ...",
      "accounting_entries": [...],
      "inventory_entries": [...],
      "gst": {{"cgst": 0, "sgst": 0, "igst": 0}},
      "total_amount": 0,
      "partner_name": "...",
      "partner_id": "..."
    }},
    ...more entries...
  ]
}}
```

Match vendors/products to EXISTING partners/products when possible.
Use correction mappings to apply learned renames.
Process EVERY data row — do not skip or summarize multiple rows into one."""

    user_msg = message or f"Analyze this uploaded document: {fname}"

    try:
        if is_image:
            b64 = base64.b64encode(content).decode("utf-8")
            resp = await _call_vision(system_prompt, user_msg, b64)
        else:
            file_text = _extract_file_text(content, fname)
            # For very large files, summarize intelligently
            if len(file_text) > 60000:
                # Take first part (headers and initial data) and last part (summary)
                file_text = file_text[:40000] + "\n\n... [MIDDLE SECTION TRUNCATED - file too large] ...\n\n" + file_text[-15000:]
            elif len(file_text) > 30000:
                file_text = file_text[:30000] + "\n... [truncated, remaining data follows same pattern]"
            resp = await _call_llm(system_prompt, f"{user_msg}\n\nFILE CONTENT:\n{file_text}", "gpt-4o")

        data = _parse_json(resp)
        entries_list = _ensure_batch(data)

        now_iso = datetime.now(timezone.utc).isoformat()
        user_name = (await db.users.find_one({"id": current_user["user_id"]}, {"name": 1, "_id": 0}) or {}).get("name", "")

        if len(entries_list) == 1:
            # Single entry
            entry = entries_list[0]
            pending_id = str(uuid.uuid4())
            await db.ai_pending_entries.insert_one({
                "_id": pending_id,
                "id": pending_id,
                "company_id": cid,
                "user_id": current_user["user_id"],
                "user_name": user_name,
                "action_type": entry.get("action_type", "unknown"),
                "entries": entry,
                "original_message": user_msg,
                "file_url": file_url,
                "file_name": fname,
                "status": "pending",
                "created_at": now_iso,
                "source": "file_upload",
            })
            return {
                "type": "preview",
                "pending_id": pending_id,
                "message": entry.get("description", f"Analyzed {fname} — review the entries below."),
                "document_type": entry.get("document_type", "unknown"),
                "confidence": entry.get("confidence", 0),
                "entries": entry,
                "file_url": file_url,
                "file_name": fname,
            }
        else:
            # Multiple entries — batch mode
            batch_ids = []
            batch_ref = str(uuid.uuid4())[:8]
            for idx, entry in enumerate(entries_list):
                pending_id = str(uuid.uuid4())
                await db.ai_pending_entries.insert_one({
                    "_id": pending_id,
                    "id": pending_id,
                    "company_id": cid,
                    "user_id": current_user["user_id"],
                    "user_name": user_name,
                    "action_type": entry.get("action_type", "unknown"),
                    "entries": entry,
                    "original_message": user_msg,
                    "file_url": file_url,
                    "file_name": fname,
                    "status": "pending",
                    "created_at": now_iso,
                    "source": "file_upload",
                    "batch_ref": batch_ref,
                    "batch_index": idx,
                })
                batch_ids.append({"pending_id": pending_id, "entries": entry})
            return {
                "type": "batch_preview",
                "batch_ref": batch_ref,
                "message": f"Analyzed {fname} — found {len(batch_ids)} entries for your review.",
                "batch": batch_ids,
                "total_count": len(batch_ids),
                "file_url": file_url,
                "file_name": fname,
            }

    except json.JSONDecodeError:
        return {"type": "error", "message": "Could not parse AI response", "raw": resp[:500] if resp else ""}
    except Exception as e:
        logger.error(f"AI upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 3. APPROVE & POST ============

@router.post("/approve")
async def approve_entry(req: ApproveRequest, current_user: dict = Depends(get_current_user)):
    """Approve and post a pending AI-generated entry (with optional edits)."""
    if current_user['role'] == UserRole.GROUND_STAFF:
        raise HTTPException(status_code=403, detail="Access denied")

    pending = await db.ai_pending_entries.find_one({"id": req.pending_id}, {"_id": 0})
    if not pending:
        raise HTTPException(status_code=404, detail="Pending entry not found")
    if pending["status"] != "pending":
        raise HTTPException(status_code=400, detail=f"Entry already {pending['status']}")

    cid = pending["company_id"]
    entries = req.entries if req.entries else pending["entries"]
    action_type = entries.get("action_type", pending.get("action_type", ""))
    results = []

    try:
        # Post accounting entries
        if entries.get("accounting_entries"):
            acc_lines = entries["accounting_entries"]
            move_data = {
                "move_type": "entry",
                "journal_id": None,
                "ref": f"AI-{pending.get('source','chat')}-{req.pending_id[:8]}",
                "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "narration": entries.get("description", "AI-generated entry"),
                "lines": [],
            }
            # Find misc journal
            misc_journal = await db.odoo_journals.find_one({"company_id": cid, "code": "MISC"}, {"_id": 0, "id": 1})
            if misc_journal:
                move_data["journal_id"] = misc_journal["id"]

            for line in acc_lines:
                acct_id = line.get("account_id")
                if not acct_id and line.get("account_code"):
                    acct = await db.odoo_accounts.find_one({"company_id": cid, "code": line["account_code"]}, {"_id": 0, "id": 1})
                    acct_id = acct["id"] if acct else None
                if acct_id:
                    move_data["lines"].append({
                        "account_id": acct_id,
                        "name": line.get("description", line.get("account_name", "")),
                        "debit": float(line.get("debit", 0)),
                        "credit": float(line.get("credit", 0)),
                    })

            if move_data["lines"] and move_data["journal_id"]:
                # Create journal entry directly in MongoDB
                move_id = str(uuid.uuid4())
                journal = await db.odoo_journals.find_one({"id": move_data["journal_id"], "company_id": cid}, {"_id": 0})
                now_iso = datetime.now(timezone.utc).isoformat()
                total_debit = sum(ln["debit"] for ln in move_data["lines"])

                # Build move lines
                move_lines = []
                for line in move_data["lines"]:
                    ml_id = str(uuid.uuid4())
                    move_lines.append({
                        "id": ml_id, "move_id": move_id,
                        "account_id": line["account_id"],
                        "name": line.get("name", ""),
                        "debit": round(line["debit"], 2),
                        "credit": round(line["credit"], 2),
                        "company_id": cid,
                    })

                move_doc = {
                    "id": move_id,
                    "name": f"AI/{move_data['ref']}",
                    "move_type": "entry",
                    "journal_id": move_data["journal_id"],
                    "journal_name": journal["name"] if journal else "Miscellaneous",
                    "ref": move_data["ref"],
                    "narration": move_data.get("narration", ""),
                    "date": move_data["date"],
                    "state": "posted",
                    "amount_total": round(total_debit, 2),
                    "company_id": cid,
                    "created_by": current_user["user_id"],
                    "created_at": now_iso,
                    "posted_at": now_iso,
                }

                await db.odoo_moves.insert_one({**move_doc, "_id": move_id})
                for ml in move_lines:
                    await db.odoo_move_lines.insert_one({**ml, "_id": ml["id"]})

                results.append({"type": "journal_entry", "id": move_id, "name": move_doc["name"]})

        # Post inventory entries
        if entries.get("inventory_entries"):
            for inv_entry in entries["inventory_entries"]:
                qty = float(inv_entry.get("quantity_change", 0))
                if qty == 0:
                    continue
                product_id = inv_entry.get("product_id")
                if not product_id and inv_entry.get("product_name"):
                    prod = await db.products.find_one({"company_id": cid, "name": {"$regex": inv_entry["product_name"], "$options": "i"}}, {"_id": 0, "id": 1})
                    product_id = prod["id"] if prod else None

                if product_id:
                    wh_id = inv_entry.get("warehouse_id")
                    if not wh_id:
                        wh = await db.stock_warehouses.find_one({"company_id": cid}, {"_id": 0, "id": 1})
                        wh_id = wh["id"] if wh else None

                    from odoo_inventory.engine import inventory_adjustment
                    adj = await inventory_adjustment(db, cid, {
                        "product_id": product_id,
                        "new_quantity": max(0, (await db.products.find_one({"id": product_id}, {"_id": 0, "quantity_on_hand": 1}) or {}).get("quantity_on_hand", 0) + qty),
                        "reason": entries.get("description", "AI adjustment"),
                    })
                    if adj:
                        results.append({"type": "inventory_adjustment", "product_id": product_id, "quantity_change": qty})

        # Update pending status
        await db.ai_pending_entries.update_one(
            {"id": req.pending_id},
            {"$set": {
                "status": "approved",
                "approved_by": current_user["user_id"],
                "approved_at": datetime.now(timezone.utc).isoformat(),
                "posted_entries": entries,
                "results": results,
            }}
        )

        # Audit trail
        user_doc = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "name": 1, "role": 1})
        await db.ai_audit_trail.insert_one({
            "id": str(uuid.uuid4()),
            "pending_id": req.pending_id,
            "company_id": cid,
            "action": "approved_and_posted",
            "action_type": action_type,
            "reviewed_by": user_doc.get("name", ""),
            "user_role": user_doc.get("role", ""),
            "results": results,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "source": pending.get("source", ""),
            "file_url": pending.get("file_url"),
        })

        return {"status": "posted", "results": results, "message": f"Successfully posted {len(results)} entries."}

    except Exception as e:
        logger.error(f"Approve error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 4. REJECT ============

@router.post("/reject")
async def reject_entry(req: RejectRequest, current_user: dict = Depends(get_current_user)):
    pending = await db.ai_pending_entries.find_one({"id": req.pending_id}, {"_id": 0})
    if not pending:
        raise HTTPException(status_code=404, detail="Not found")
    await db.ai_pending_entries.update_one(
        {"id": req.pending_id},
        {"$set": {"status": "rejected", "rejected_by": current_user["user_id"],
                  "rejected_at": datetime.now(timezone.utc).isoformat()}}
    )
    await db.ai_audit_trail.insert_one({
        "id": str(uuid.uuid4()), "pending_id": req.pending_id,
        "company_id": pending["company_id"], "action": "rejected",
        "reviewed_by": current_user["user_id"],
        "timestamp": datetime.now(timezone.utc).isoformat(),
    })
    return {"status": "rejected"}


# ============ 4b. BATCH APPROVE ============

@router.post("/batch-approve")
async def batch_approve(req: BatchApproveRequest, current_user: dict = Depends(get_current_user)):
    """Approve multiple pending entries at once."""
    if current_user['role'] == UserRole.GROUND_STAFF:
        raise HTTPException(status_code=403, detail="Access denied")

    all_results = []
    errors = []
    for pid in req.pending_ids:
        try:
            # Reuse single approve logic
            inner_req = ApproveRequest(pending_id=pid)
            result = await approve_entry(inner_req, current_user)
            all_results.append({"pending_id": pid, "status": "posted", "results": result.get("results", [])})
        except Exception as e:
            errors.append({"pending_id": pid, "error": str(e)})

    return {
        "status": "batch_posted",
        "total_approved": len(all_results),
        "total_errors": len(errors),
        "results": all_results,
        "errors": errors,
        "message": f"Successfully posted {len(all_results)} of {len(req.pending_ids)} entries.",
    }


# ============ 4c. BATCH REJECT ============

@router.post("/batch-reject")
async def batch_reject(req: BatchApproveRequest, current_user: dict = Depends(get_current_user)):
    """Reject multiple pending entries at once."""
    count = 0
    for pid in req.pending_ids:
        try:
            inner_req = RejectRequest(pending_id=pid)
            await reject_entry(inner_req, current_user)
            count += 1
        except Exception:
            pass
    return {"status": "batch_rejected", "total_rejected": count}


# ============ 5. HISTORY ============

@router.get("/history")
async def get_history(company_id: Optional[str] = None, limit: int = 50,
                      current_user: dict = Depends(get_current_user)):
    cid = await _cid(current_user, company_id)
    entries = await db.ai_pending_entries.find(
        {"company_id": cid}, {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    return entries


# ============ 6. AUDIT TRAIL ============

@router.get("/audit-trail")
async def get_audit_trail(company_id: Optional[str] = None, limit: int = 100,
                          current_user: dict = Depends(get_current_user)):
    cid = await _cid(current_user, company_id)
    trail = await db.ai_audit_trail.find(
        {"company_id": cid}, {"_id": 0}
    ).sort("timestamp", -1).limit(limit).to_list(limit)
    return trail


# ============ 7. SMART LEARNING — CORRECTION MAPPINGS ============

@router.post("/learn")
async def save_correction(mapping: CorrectionMapping, current_user: dict = Depends(get_current_user)):
    """Save a user correction so AI learns for future documents."""
    cid = await _cid(current_user, mapping.company_id)
    doc = {
        "id": str(uuid.uuid4()),
        "company_id": cid,
        "original": mapping.original,
        "corrected": mapping.corrected,
        "field": mapping.field,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.ai_correction_mappings.update_one(
        {"company_id": cid, "original": mapping.original, "field": mapping.field},
        {"$set": doc}, upsert=True,
    )
    return {"status": "saved", "mapping": {k: v for k, v in doc.items()}}


@router.get("/mappings")
async def get_mappings(company_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    cid = await _cid(current_user, company_id)
    return await db.ai_correction_mappings.find({"company_id": cid}, {"_id": 0}).to_list(200)


@router.delete("/mappings/{mapping_id}")
async def delete_mapping(mapping_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a correction mapping."""
    result = await db.ai_correction_mappings.delete_one({"id": mapping_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Mapping not found")
    return {"status": "deleted"}


# ============ 8. AUDIT TRAIL STATS ============

@router.get("/audit-stats")
async def get_audit_stats(company_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get aggregated audit trail stats."""
    cid = await _cid(current_user, company_id)
    total_approved = await db.ai_audit_trail.count_documents({"company_id": cid, "action": "approved_and_posted"})
    total_rejected = await db.ai_audit_trail.count_documents({"company_id": cid, "action": "rejected"})
    total_pending = await db.ai_pending_entries.count_documents({"company_id": cid, "status": "pending"})
    return {
        "total_approved": total_approved,
        "total_rejected": total_rejected,
        "total_pending": total_pending,
        "total_actions": total_approved + total_rejected,
    }
